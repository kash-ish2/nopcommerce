import openpyxl


class ExcelUtils:

    # Get total number of rows
    @staticmethod
    def get_row_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row


    # Get total number of columns
    @staticmethod
    def get_column_count(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_column


    # Read data from specific cell
    @staticmethod
    def read_data(file_path, sheet_name, row, column):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value


    # Write data to specific cell
    @staticmethod
    def write_data(file_path, sheet_name, row, column, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(file_path)


    # Get all data from sheet (returns list of lists)
    @staticmethod
    def get_all_data(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        data = []

        for row in sheet.iter_rows(values_only=True):
            data.append(list(row))

        return data


    # Get data from entire row
    @staticmethod
    def get_row_data(file_path, sheet_name, row):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        row_data = []
        for col in range(1, sheet.max_column + 1):
            row_data.append(sheet.cell(row=row, column=col).value)

        return row_data


    # Get data from entire column
    @staticmethod
    def get_column_data(file_path, sheet_name, column):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        col_data = []
        for row in range(1, sheet.max_row + 1):
            col_data.append(sheet.cell(row=row, column=column).value)

        return col_data
